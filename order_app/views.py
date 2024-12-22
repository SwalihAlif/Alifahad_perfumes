from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from order_app.models import Order, OrderItems, CancelledOrder
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.db import transaction
from cart_app.models import Cart, CartItem
from order_app.models import Order, OrderItems
from user_profile_app.models import UserProfile
from product_app.models import Product, Variant
from coupon_app.models import Coupon
from wallet_app.models import Wallet, WalletTransactions
from validator_app import views
from django.db.models import Sum, F
from django.db.models.functions import TruncMonth
from reportlab.pdfgen import canvas
from io import BytesIO
import razorpay
import json
import logging
from django.conf import settings
from decimal import Decimal
from django.utils import timezone
from django.utils.timezone import now
from django.shortcuts import render
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Sum, Count

from django.conf import settings
import razorpay

razorpay_client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))

@login_required(login_url='login')
def checkout(request):
    user = request.user
    addresses = UserProfile.objects.filter(user=user, delete_address=False)
    wallet, created = Wallet.objects.get_or_create(user=user)
    available_coupons = Coupon.objects.filter(usage_limit__gt=0, active=True)

    try:
        cart = Cart.objects.get(user_id=user)
        cart_items = CartItem.objects.filter(cart=cart, variant__product__is_listed=True)

        if not cart_items.exists():
            return JsonResponse({"error": "Your cart is empty. Please add items to checkout."}, status=400)

        total = Decimal('0.00')
        discount = Decimal('0.00') # Initialize discount variable
        for item in cart_items:
            price = item.variant.product_price_after()
            quantity = item.quantity
            item.sub_total = price * quantity
            total += item.sub_total

        shipping = Decimal('150.00') if total < Decimal('1000.00') else Decimal('0.00')
        total += shipping

        if request.method == 'POST':
            try:
                data = json.loads(request.body.decode("utf-8"))
                address_id = data.get("address_id")
                payment_method = data.get("payment_method")
                coupon_code = data.get("coupon_code")

                if not address_id or not payment_method:
                    return JsonResponse({"error": "Address or payment method is not provided."}, status=400)

                address = UserProfile.objects.get(id=address_id, user=user)

                # Coupon logic
                if coupon_code:
                    try:
                        coupon = Coupon.objects.get(code=coupon_code)
                        if coupon.active and (not coupon.valid_until or coupon.valid_until >= timezone.now()):
                            if total >= coupon.min_purchase_amount:
                                if coupon.discount_type == 'percentage':
                                    discount = total * (coupon.discount_amount / Decimal('100'))
                                elif coupon.discount_type == 'fixed':
                                    discount = coupon.discount_amount
                                else:
                                    discount = Decimal('0.00')

                                total -= discount
                                coupon.usage_limit -= 1
                                if coupon.usage_limit <= 0:
                                    coupon.active = False
                                coupon.save()
                            else:
                                return JsonResponse({'error': 'Total does not meet the minimum purchase amount for this coupon'}, status=400)
                        else:
                            return JsonResponse({'error': 'Coupon is not active or has expired'}, status=400)
                    except Coupon.DoesNotExist:
                        return JsonResponse({'error': 'Invalid coupon code'}, status=400)
                    
                # Ensure total is never less than zero 
                total = max(total, Decimal('0.00'))

                if payment_method == "COD" and total > 1000:
                    return JsonResponse({"error": "Order above 1000 is not available for COD."}, status=400)

                if payment_method == 'wallet' and total > wallet.balance:
                    return JsonResponse({"error": "Not enough money in wallet. Try another payment method."}, status=400)

                # Check stock availability for all items
                for item in cart_items:
                    variant = Variant.objects.get(product=item.variant.product, size=item.variant.size)
                    if variant.stock < item.quantity:
                        return JsonResponse({
                            "error": f"Not enough stock for {item.variant.product.product_name} (Size: {item.variant.size})."
                        }, status=400)
                    print(f"This is ------------------------------------------ varient{variant}")
                    
                

                # Start order creation in a transaction to ensure atomicity
                with transaction.atomic():
                    order = Order.objects.create(
                        user_id=user,
                        address=address,
                        payment_method=payment_method,
                        coupon_discount = discount,
                        total_amount=total,
                        status="Order Pending"
                    )
                    print(f"this is -------------------------------------- order{order}")

                    if payment_method == 'razorpay':
                        # Deduct stock for Razorpay payment method
                        for item in cart_items:
                            variant = Variant.objects.get(product=item.variant.product, size=item.variant.size)
                            variant.stock -= item.quantity
                            variant.save()

                            price = item.variant.product_price_after()

                            OrderItems.objects.create(
                                order=order,
                                product_added=item.variant,
                                quantity=item.quantity,
                                final_product_price=price,
                                size=variant,
                            )

                        cart_items.delete()

                        # Razorpay order creation
                        currency = 'INR'
                        amount = int(total * 100)  # Convert to paise

                        razorpay_order = razorpay_client.order.create({
                            'amount': amount,
                            'currency': currency,
                            'payment_capture': '1'
                        })

                        order.payment_online_id = razorpay_order['id']
                        order.save()

                        payment_data = {
                            'order_id': razorpay_order['id'],
                            'amount': amount,
                            'currency': currency,
                            'key': settings.RAZORPAY_KEY_ID,
                            'success': True
                        }
                        return JsonResponse(payment_data)

                    elif payment_method == "COD":
                        order.payment_status = "pending"
                        order.save()

                    
                    
                    elif payment_method == "wallet":
                        total = float(total)
                        if wallet.balance >= total:
                            order.payment_status = "completed"
                            order.save()
                            wallet.balance -= total
                            wallet.save()

                            WalletTransactions.objects.create(
                                user=request.user,
                                order=order,
                                wallet=wallet,
                                amount=total,
                                transaction_type=WalletTransactions.SPENT,
                                description=f"Spent for {order.serial_number} "
                            )
                        else:
                            return JsonResponse({"error": "Not enough funds in wallet."}, status=400)

                    # Deduct stock for COD and Wallet payment methods
                    for item in cart_items:
                        variant = Variant.objects.get(product=item.variant.product, size=item.variant.size)
                        variant.stock -= item.quantity
                        variant.save()

                        price = item.variant.product_price_after()
                        print(f"This is ---------------------------------------------- price{price}")

                        OrderItems.objects.create(
                            order=order,
                            product_added=item.variant,
                            quantity=item.quantity,
                            final_product_price=price,
                            size=variant,
                        )

                    cart_items.delete()

                    return JsonResponse({"success": "Order placed successfully!", "order_id": order.serial_number}, status=200)

            except json.JSONDecodeError:
                return JsonResponse({"error": "Invalid request format."}, status=400)

    except Cart.DoesNotExist:
        return JsonResponse({"error": "Cart does not exist."}, status=400)

    context = {
        'addresses': addresses,
        'cart_items': cart_items,
        'total': total,
        'delivery_charge': shipping,
        'wallet': wallet,
        'available_coupons': available_coupons,
        
    }
    return render(request, 'user/checkout.html', context)

      
#--------------------------- Apply coupon -----------------------------------------------------------------------


@login_required(login_url='login')
@csrf_exempt
def apply_coupon(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        coupon_code = data.get('coupon_code')
        total_amount_excluding_shipping = data.get('total_amount_excluding_shipping')
        
        if not coupon_code:
            return JsonResponse({'valid': False, 'error': 'No coupon code provided'})

        try:
            coupon = Coupon.objects.get(code=coupon_code)
            
            if not total_amount_excluding_shipping:
                return JsonResponse({'valid': False, 'error': 'Total amount excluding shipping not provided'})
            
            # Check if the total amount excluding shipping is greater than or equal to the minimum purchase amount
            if total_amount_excluding_shipping < coupon.min_purchase_amount:
                return JsonResponse({'valid': False, 'error': f'Minimum purchase amount not met. Minimum required is ₹{coupon.min_purchase_amount}.'})
            
            if coupon.active and (not coupon.valid_until or coupon.valid_until >= timezone.now()):
                return JsonResponse({'valid': True, 'discount': coupon.discount_amount, 'message': 'Coupon applied successfully'})
            else:
                return JsonResponse({'valid': False, 'error': 'Coupon is not active or has expired'})
        except Coupon.DoesNotExist:
            return JsonResponse({'valid': False, 'error': 'Invalid coupon code'})
        
    return JsonResponse({'valid': False, 'error': 'Invalid request method'})


#------------------------ Order success --------------------------------------------------------------------------------

@never_cache
@login_required(login_url='login')
def order_success(request):
    try:
        # Get the latest order for the user
        latest_order = Order.objects.filter(user_id=request.user).latest('order_date')

        # Get order items for the latest order
        order_items = OrderItems.objects.filter(order=latest_order)

        context = {
            'order': latest_order,
            'order_items': order_items,
        }
        return render(request, 'user/order_confirm.html', context)

    except Order.DoesNotExist:
        return redirect('user:checkout')



    
#---------------------------- User invoice dowload ------------------------------------------------------------------
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.colors import navy, red, black
from .models import Order, OrderItems, UserProfile

@login_required(login_url='login')
def download_invoice(request, order_id):
    try:
        # Get the specified order for the user
        order = Order.objects.get(serial_number=order_id, user_id=request.user)
        order_items = OrderItems.objects.filter(order=order)

        # Create a file-like buffer to receive PDF data
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter

        # Draw a border around the page
        p.setStrokeColorRGB(0, 0, 0)
        p.setLineWidth(1)
        p.rect(10, 10, width - 20, height - 20)

        
        # Add the website logo with specified dimensions
        logo_path = 'static/user/images/icons/dis.png'
        logo_width = 4 / 2.54 * inch  # Convert cm to inches (1 cm = 0.393701 inches)
        logo_height = 1 / 2.54 * inch  # Convert cm to inches (1 cm = 0.393701 inches)
        p.drawImage(logo_path, 40, height - (0.5 / 2.54 * inch) - logo_height, width=logo_width, height=logo_height)



        # Draw the title with color
        p.setFont("Helvetica-Bold", 16)
        p.setFillColor(navy)
        p.drawString(200, height - 100, "Invoice")

        # Order details with colors
        p.setFont("Helvetica", 12)
        p.setFillColor(black)
        p.drawString(100, height - 140, f"Order ID: {order.serial_number}")
        p.setFillColor(red)
        p.drawString(100, height - 160, f"Order Date: {order.order_date}")
        p.setFillColor(black)
        p.drawString(100, height - 180, f"Payment Method: {order.payment_method}")
        p.setFillColor(red)
        p.drawString(100, height - 200, f"Payment Status: {order.payment_status}")
        p.setFillColor(black)
        p.drawString(100, height - 220, f"Total Amount: Rs. {order.total_amount}")
        p.setFillColor(red)
        p.drawString(100, height - 240, f"Discount: Rs.{order.coupon_discount or 0}")

        # Shipping address
        p.setFillColor(navy)
        p.drawString(100, height - 270, "Shipping Address:")
        p.setFillColor(black)
        p.drawString(120, height - 290, f"{order.address.full_name}")
        p.drawString(120, height - 310, f"{order.address.area}, {order.address.city}")
        p.drawString(120, height - 330, f"{order.address.state} - {order.address.pincode}")
        p.drawString(120, height - 350, f"Phone: {order.address.phone_no}")

        # Order items
        p.setFillColor(navy)
        p.drawString(100, height - 380, "Order Items:")
        y = height - 410
        for item in order_items:
            p.setFillColor(black)
            p.drawString(120, y, f"Product: {item.product_added.product.product_name}")
            p.setFillColor(red)
            p.drawString(120, y - 20, f"Size: {item.size}")
            p.setFillColor(black)
            p.drawString(120, y - 40, f"Quantity: {item.quantity}")
            p.setFillColor(red)
            p.drawString(120, y - 60, f"Price: ₹{item.final_product_price}")
            y -= 80

        # Add a "Thank You" message
        p.setFillColor(navy)
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, y - 20, "Thank You for Your Order!")

        # Close the PDF object cleanly, and we're done.
        p.showPage()
        p.save()

        # Get the value of the BytesIO buffer and write it to the response.
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')

    except Order.DoesNotExist:
        return redirect('user:order_success')  # Redirect to order success page if order does not exist


#---------------------------- Razorpay payment handler ------------------------------------------------------------------

logger = logging.getLogger(__name__)

razorpay_client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))

@csrf_exempt
@require_POST
def payment_handler(request):
    try:
        payment_id = request.POST.get('razorpay_payment_id')
        order_id = request.POST.get('razorpay_order_id')
        signature = request.POST.get('razorpay_signature')

        logger.info(f"Received payment verification request - Order ID: {order_id}, Payment ID: {payment_id}")

        if not all([payment_id, order_id, signature]):
            logger.error(f"Missing payment parameters: payment_id={bool(payment_id)}, order_id={bool(order_id)}, signature={bool(signature)}")
            return JsonResponse({
                'error': 'Missing required payment parameters',
                'received_params': {
                    'payment_id': payment_id,
                    'order_id': order_id,
                    'signature': signature
                }
            }, status=400)

        params_dict = {
            'razorpay_payment_id': payment_id,
            'razorpay_order_id': order_id,
            'razorpay_signature': signature
        }

        try:
            razorpay_client.utility.verify_payment_signature(params_dict)

            try:
                order = Order.objects.get(payment_online_id=order_id)
                payment = razorpay_client.payment.fetch(payment_id)

                if int(payment['amount']) != int(order.total_amount * 100):
                    logger.error(f"Payment amount mismatch - Expected: {int(order.total_amount * 100)}, Received: {payment['amount']}")
                    return JsonResponse({
                        'error': 'Payment amount does not match order amount'
                    }, status=400)

                order.payment_id = payment_id  # Add this field to the model
                order.payment_status = 'Completed'
                order.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Payment successful',
                    'order_id': order_id,
                    'order__id': order.serial_number
                })

            except Order.DoesNotExist:
                logger.error(f"Order not found: {order_id}")
                return JsonResponse({
                    'error': 'Order not found'
                }, status=400)
        except razorpay.errors.SignatureVerificationError as e:
            logger.error(f"Signature verification failed: {str(e)}")
            return JsonResponse({
                'error': 'Invalid payment signature'
            }, status=400)

    except Exception as e:
        logger.error(f"Payment processing error: {str(e)}")
        return JsonResponse({
            'error': 'Payment processing failed',
            'details': str(e)
        }, status=500)

    
#---------------------------- all orders user side --------------------------------------------------------------------


# Initialize logger and Razorpay client

logger = logging.getLogger(__name__)
razorpay_client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))

@csrf_exempt
@require_POST
def retry_payment(request):
    try:
        # Parse the JSON data from the request body
        data = json.loads(request.body.decode("utf-8"))
        order_id = data.get('orderId')
        user = request.user

        # Retrieve the order and check if it belongs to the user and payment status is pending
        order = Order.objects.get(serial_number=order_id, user_id=user)
        if not order or order.payment_status != "Pending":
            return JsonResponse({
                'val': False,
                'msg': "Invalid order or the payment was not failed",
            }, status=400)

        logger.info(f"Retrying payment for order: {order}")
        print("hellooooooooooooooooo------------------------------------------------------------retry")


        # Create a new Razorpay order
        razorpay_order = razorpay_client.order.create({
            'amount': int(order.total_amount * 100),  # Amount in paise
            'currency': 'INR',
            'receipt': str(order.serial_number),
            'notes': {
                'orderId': str(order.serial_number),
            },
        })

        order.payment_status = 'Completed'
        order.save()

        # Return the new Razorpay order details
        return JsonResponse({
            'val': True,
            'orderId': razorpay_order['id'],
            'amount': razorpay_order['amount'],
            'key': settings.RAZORPAY_KEY_ID,
        }, status=200)

    except Order.DoesNotExist:
        logger.error(f"Order not found: {order_id}")
        return JsonResponse({
            'val': False,
            'msg': "Order not found",
        }, status=400)
    except Exception as err:
        logger.error(f"Error retrying payment: {err}")
        return JsonResponse({
            'val': False,
            'msg': "Payment retry failed",
            'error': str(err),
        }, status=500)


#---------------------------- all orders user side --------------------------------------------------------------------

@login_required(login_url='login')
def all_orders(request):
    try:
        orders = Order.objects.filter(user_id=request.user.id).order_by('-order_date')
        
        paginator = Paginator(orders, 10)  
        
        # Get the current page number from the query parameters
        page_number = request.GET.get('page')
        
        try:
            paginated_orders = paginator.page(page_number)
        except PageNotAnInteger:
            # If page is not an integer, deliver the first page
            paginated_orders = paginator.page(1)
        except EmptyPage:
            # If page is out of range, deliver the last page
            paginated_orders = paginator.page(paginator.num_pages)
        
        return render(request, 'user/all_orders.html', {
            'orders': paginated_orders,
        })
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('home')


    

#---------------------------- Order details ------------------------------------------------------------------

from django.shortcuts import render, get_object_or_404
from .models import Order

@login_required(login_url='login')
def order_detail(request, serial_number):
    order = get_object_or_404(Order, serial_number=serial_number, user_id=request.user)
    order_item = OrderItems.objects.filter(order=order)
    return render(request, 'user/order_detail.html', {'order': order, 'order_item': order_item})


#---------------------------- Cancel order ------------------------------------------------------------------


@login_required(login_url='login')
def cancel_order(request, order_id):
    try:
        order = get_object_or_404(Order, user_id=request.user, serial_number=order_id)

        order_items = OrderItems.objects.filter(order=order)
        total_refund = Decimal('0.00')  # Initialize the total refund amount

        for item in order_items:
            if item.status == "Pending":
                item.status = 'Cancelled'
                item.save()

                CancelledOrder.objects.create(
                    ordered_item=item,
                    user=request.user,
                    reason=request.POST.get('reason', 'No reason provided')
                )
                total_refund += item.final_product_price * item.quantity  # Add the item's price to the total refund

        # Update order status if all items are cancelled
        if all(item.status == 'Cancelled' for item in order_items):
            order.status = 'Cancelled'
            order.save()

        # Refund the amount to the user's wallet
        wallet, created = Wallet.objects.get_or_create(user=request.user)
        wallet.balance += float(total_refund)
        wallet.save()

        # Log the wallet transaction
        WalletTransactions.objects.create(
            user=request.user,
            order=order,
            wallet=wallet,
            amount=total_refund,
            transaction_type=WalletTransactions.REFUND,
            description=f"Refund for cancelled order {order.serial_number}"
        )

        messages.success(request, "Your order has been cancelled and the amount has been refunded to your wallet.")
        return redirect('order:all_orders')

    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('order:all_orders')


#--------------------------- Admin order management ---------------------------------------------------------------


@staff_member_required
def order_management(request):

    orders_list = Order.objects.prefetch_related('order_all').order_by('-order_date')
   

    paginator = Paginator(orders_list, 6)
    page_number = request.GET.get('page')
    page_orders = paginator.get_page(page_number)

    context = {
        'orders': page_orders
    }
    return render(request, 'admin/order_management.html', context)

#--------------------------- Admin order details view -------------------------------------------------------------

def order_detail_view(request, order_serial_number):
    order = get_object_or_404(Order, serial_number=order_serial_number)
    order_items = OrderItems.objects.filter(order=order)
    address = order.address
    
    context = {
        "order": order,
        "order_items": order_items,
        "address": address,
    }
    
    return render(request, 'admin/order_detail.html', context)

#------------------------------ admin order status update ----------------------------------------------------------


@staff_member_required
def change_order_status(request, serial_number):
    order = get_object_or_404(Order, serial_number=serial_number)

    if request.method == 'POST':
        status = request.POST.get('ostatus')
        order.status = status

        if status == 'Delivered':
            order.payment_status = 'completed'
        order.save()
        messages.success(request, f"Order {order.serial_number} status is changed to {status}")

        total_refund = Decimal('0.00')  # Initialize the total refund amount

        for item in order.order_all.all():
            item.status = status  # Update the status of each order item to match the order status
            item.save()
            if status == 'Cancelled':
                total_refund += item.final_product_price * item.quantity  # Add the item's price to the total refund

        if status == 'Cancelled':
            # Refund the amount to the user's wallet
            wallet, created = Wallet.objects.get_or_create(user=order.user_id)
            wallet.balance += float(total_refund)
            wallet.save()

            # Log the wallet transaction
            WalletTransactions.objects.create(
                user=order.user_id,
                order=order,
                wallet=wallet,
                amount=total_refund,
                transaction_type=WalletTransactions.REFUND,
                description=f"Refund for cancelled order {order.serial_number}"
            )

            messages.success(request, f"All items of Order {order.serial_number} have been marked as cancelled and the amount has been refunded to the wallet.")
    
    return redirect('order:order_detail_view', order_serial_number=serial_number)




#------------------------------ admin order status update ----------------------------------------------------------
@staff_member_required
def change_order_item_status(request, oid):
    order_item = OrderItems.objects.get(id=oid)

    if request.method == 'POST':
        status = request.POST.get('status')
        print(f"order item status--------------------------------------------- {status}")
        order_item.status = status
        order_id = order_item.order
        order_item.save()
        messages.success(request, f"Order item {order_item}'s status changed to {status}")
        print(f"this is {order_item}")

        order = order_item.order
        all_same_status = all(item.status == status for item in order.order_all.all())

        if all_same_status:
            order.status = status
            order.save()
            messages.success(request, f"Order {order.serial_number} status changed to {status}")

    return redirect('order:order_detail_view', order_id)

#-----------------------------User Return request ----------------------------------------------------------------------------------
@require_POST
def request_item_return(request, item_id):
    order_item = get_object_or_404(OrderItems, id=item_id)

    if order_item.status == 'Delivered':
        order_item.item_return_requested = True
        order_item.status = 'Requested'
        order_item.save()
        messages.success(request, "Return request submitted successfully.")
    else:
        messages.error(request, "Return request can only be submitted delivered items.")

    return redirect('order:order_detail', serial_number=order_item.order.serial_number)



#----------------------------- Admin approve return request ----------------------------------------------------------------------------------
@staff_member_required
@require_POST
def approve_item_return(request, item_id):
    order_item = get_object_or_404(OrderItems, id=item_id)

    if order_item.item_return_requested: 
        order_item.status = 'Returned'
        order_item.item_return_requested = False
        order_item.save()

        order = order_item.order
        order.total_amount -= order_item.final_product_price
        # order.status = 'Returned'
        
        order.save()

        wallet = get_object_or_404(Wallet, user=order.user_id)
        retuned_amount = float(order_item.final_product_price)
        wallet.balance += retuned_amount
        wallet.save()

        WalletTransactions.objects.create(
            user=request.user,
            order=order,
            wallet=wallet,
            amount=retuned_amount,
            transaction_type=WalletTransactions.REFUND,
            description=f"Refund of the {order.serial_number} "
        )

        all_items_returned = all(item.status == 'Returned' for item in order.order_all.all())
        if all_items_returned:
            order.status = 'Returned'
            order.payment_status = 'Refunded'
            order.save()

        messages.success(request, f"Return approved and amount refunded to the {wallet.user.username}'s wallet.")
    else:
        messages.error(request, "No return request for this item")

    return redirect('order:order_detail_view', order_serial_number=order_item.order.serial_number)
#----------------------------- Sales report ----------------------------------------------------------------------------------


@staff_member_required
def dashboard_view(request):

    

# Filter orders with the status 'Delivered'
    delivered_orders = Order.objects.filter(status='Delivered')

# Aggregate total sales and count of delivered orders
    total_sales = delivered_orders.aggregate(total=Sum('total_amount'))['total'] or 0
    sales_count = delivered_orders.aggregate(count=Count('serial_number'))['count'] or 0
    

    top_category = OrderItems.objects.values('product_added__product__category__category_name', 'product_added__product__category__category_unit', 'product_added__product__category__category_image')\
        .annotate(total_quantity_sold=Sum('quantity'), total_revenue=Sum(F('quantity') * F('final_product_price')))\
        .order_by('-total_quantity_sold')
    

    top_products = OrderItems.objects.values('product_added__product__product_name', 'product_added__product__image_1', 'product_added__product__category__category_unit', 'product_added__size', 'product_added__stock')\
        .annotate(total_quantity_sold=Sum('quantity'), total_revenue=Sum(F('quantity') * F('final_product_price')))\
        .order_by('-total_quantity_sold')
    
    paginator = Paginator(top_products, 10)  
    page_number = request.GET.get('page') 
    page_products = paginator.get_page(page_number)

    top_customers = Order.objects.values('user_id__username')\
        .annotate(total_orders=Count('serial_number'), total_spend=Sum('total_amount'))\
        .order_by('-total_spend')[:5]  

    orders_by_pincode = Order.objects.values('address__pincode', 'address__state', 'address__city')\
        .annotate(total_orders=Count('serial_number'))\
        .order_by('-total_orders')[:5] 

    

    context = {
        'total_sales': total_sales,
        'sales_count': sales_count,
        'top_category': top_category,
        'top_products': page_products,
        'top_customers': top_customers,
        'orders_by_pincode': orders_by_pincode,
    }
    return render(request, 'admin/dashboard.html', context)


#----------------------------- top categories ------------------------------------------------------------------------
@staff_member_required
def top_categories(request):
   
    top_category = OrderItems.objects.values('product_added__product__category__category_name', 'product_added__product__category__category_unit', 'product_added__product__category__category_image')\
        .annotate(total_quantity_sold=Sum('quantity'), total_revenue=Sum(F('quantity') * F('final_product_price')))\
        .order_by('-total_quantity_sold')
    
    context = {
        'top_category': top_category,
    }
    return render(request, 'admin/top_categories.html', context)
#----------------------------- top products ------------------------------------------------------------------------
@staff_member_required
def top_products(request):
    
    top_products = OrderItems.objects.values('product_added__product__product_name', 'product_added__product__image_1', 'product_added__product__category__category_unit', 'product_added__size', 'product_added__stock')\
        .annotate(total_quantity_sold=Sum('quantity'), total_revenue=Sum(F('quantity') * F('final_product_price')))\
        .order_by('-total_quantity_sold')
    
    paginator = Paginator(top_products, 10)  
    page_number = request.GET.get('page') 
    page_products = paginator.get_page(page_number)
 
    context = {
        'top_products': page_products,
    }
    return render(request, 'admin/top_products.html', context)
#----------------------------- top customers ------------------------------------------------------------------------
@staff_member_required
def top_customers(request):
    
    top_customers = Order.objects.values('user_id__username')\
        .annotate(total_orders=Count('serial_number'), total_spend=Sum('total_amount'))\
        .order_by('-total_spend')[:5]  

    context = {
        'top_customers': top_customers,
    }
    return render(request, 'admin/top_customers.html', context)
#----------------------------- top pincodes ------------------------------------------------------------------------

@staff_member_required
def top_pincodes(request):

    orders_by_pincode = Order.objects.values('address__pincode', 'address__state', 'address__city')\
        .annotate(total_orders=Count('serial_number'))\
        .order_by('-total_orders')[:5] 

    

    context = {
        'orders_by_pincode': orders_by_pincode,
    }
    return render(request, 'admin/top_pincodes.html', context)

#----------------------------- monthly orders graph ------------------------------------------------------------------------


from django.http import JsonResponse
from django.db.models import Sum
from .models import Order

def get_monthly_orders(request, year):
    orders = Order.objects.filter(order_date__year=year)\
                          .annotate(month=TruncMonth('order_date'))\
                          .values('month')\
                          .annotate(total_orders=Count('serial_number'), total_revenue=Sum('total_amount'))\
                          .order_by('month')
    data = {
        'months': [order['month'].strftime('%B') for order in orders],
        'total_orders': [order['total_orders'] for order in orders],
        'total_revenue': [order['total_revenue'] for order in orders]
    }
    return JsonResponse(data)

#----------------------------- add address in checkout ------------------------------------------------------------------------



def get_top_selling_categories(request):
    # Aggregate total sales for each category
    data = (
        OrderItems.objects
        .values('product_added__product__category__category_name')
        .annotate(total_sales=Sum('quantity'))
        .order_by('-total_sales')[:5]  # Top 5 categories
    )

    # Format data for the chart
    categories = [item['product_added__product__category__category_name'] for item in data]
    sales = [item['total_sales'] for item in data]

    return JsonResponse({'categories': categories, 'sales': sales})

#----------------------------- report download as pdf ------------------------------------------------------------------------

from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing, Rect
from reportlab.graphics.charts.piecharts import Pie
from reportlab.lib import colors
from io import BytesIO
from django.db.models import Sum, Count, F

def generate_pdf_report(request):
    # Create a file-like buffer to receive PDF data
    buffer = BytesIO()

    # Create the PDF object, using the buffer as its "file."
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Draw a border around the page
    p.setStrokeColorRGB(0, 0, 0)
    p.setLineWidth(1)
    p.rect(10, 10, width - 20, height - 20)

    # Draw the title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, height - 40, "Sales Report")

    # Total sales and sales count
    total_sales = Order.objects.aggregate(total=Sum('total_amount'))['total'] or 0
    sales_count = Order.objects.aggregate(count=Count('serial_number'))['count'] or 0

    p.setFont("Helvetica", 12)
    p.drawString(100, height - 70, f"Total Sales: ₹{total_sales}")
    p.drawString(100, height - 90, f"Sales Count: {sales_count}")

    # Top-selling categories
    top_category = OrderItems.objects.values('product_added__product__category__category_name')\
        .annotate(total_quantity_sold=Sum('quantity'), total_revenue=Sum(F('quantity') * F('final_product_price')))\
        .order_by('-total_quantity_sold')[:5]

    p.drawString(100, height - 130, "Top-Selling Categories:")

    y = height - 150
    for category in top_category:
        p.drawString(120, y, f"{category['product_added__product__category__category_name']}: {category['total_quantity_sold']} units sold, ₹{category['total_revenue']} revenue")
        y -= 20

    # Top customers
    top_customers = Order.objects.values('user_id__username')\
        .annotate(total_orders=Count('serial_number'), total_spend=Sum('total_amount'))\
        .order_by('-total_spend')[:5]

    p.drawString(100, y - 20, "Top Customers:")
    y -= 40
    for customer in top_customers:
        p.drawString(120, y, f"{customer['user_id__username']}: {customer['total_orders']} orders, ₹{customer['total_spend']} spent")
        y -= 20

    # Orders by pincode
    orders_by_pincode = Order.objects.values('address__pincode', 'address__state', 'address__city')\
        .annotate(total_orders=Count('serial_number'))\
        .order_by('-total_orders')[:5]

    p.drawString(100, y - 20, "Orders by Pincode:")
    y -= 40
    for order in orders_by_pincode:
        p.drawString(120, y, f"Pincode: {order['address__pincode']}, State: {order['address__state']}, District: {order['address__city']}, Total Orders: {order['total_orders']}")
        y -= 20

    # Adding space before pie chart
    y -= 40

    # Adding a pie chart for Top-Selling Categories
    d = Drawing(200, 200)
    pie = Pie()
    pie.x = 50
    pie.y = 50
    pie.width = 100
    pie.height = 100
    pie.data = [category['total_quantity_sold'] for category in top_category]
    pie.labels = [category['product_added__product__category__category_name'] for category in top_category]
    pie.slices.strokeWidth = 0.5
    pie.slices[3].popout = 10
    pie.slices[3].strokeWidth = 2
    pie.slices[3].strokeDashArray = [2, 2]
    pie.slices[3].labelRadius = 1.75
    pie.slices[3].fontColor = colors.red

    d.add(pie)
    d.drawOn(p, 300, y - 150)

    # Close the PDF object cleanly, and we're done.
    p.showPage()
    p.save()

    # Get the value of the BytesIO buffer and write it to the response.
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')


#----------------------------- add address in checkout ------------------------------------------------------------------------
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.db.models import Sum, Count, F

def generate_excel_report(request):
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill_color = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Add a title
    ws.merge_cells('A1:E1')
    ws['A1'] = "Sales Report"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Add headers for total sales and sales count
    ws.append(["Total Sales", "Sales Count"])
    total_sales = Order.objects.aggregate(total=Sum('total_amount'))['total'] or 0
    sales_count = Order.objects.aggregate(count=Count('serial_number'))['count'] or 0
    ws.append([total_sales, sales_count])

    # Style the header for total sales and sales count
    for cell in ws["A2:E2"][0]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = fill_color
        cell.border = thin_border

    # Add headers for top-selling categories
    ws.append([])  # Add an empty row for spacing
    ws.merge_cells('A4:E4')
    ws['A4'] = "Top-Selling Categories"
    ws['A4'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A4'].alignment = Alignment(horizontal="center")
    ws['A4'].fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
    ws.append(["Category Name", "Quantity Sold", "Total Revenue"])

    top_category = OrderItems.objects.values('product_added__product__category__category_name')\
        .annotate(total_quantity_sold=Sum('quantity'), total_revenue=Sum(F('quantity') * F('final_product_price')))\
        .order_by('-total_quantity_sold')[:5]

    for category in top_category:
        ws.append([
            category['product_added__product__category__category_name'],
            category['total_quantity_sold'],
            category['total_revenue']
        ])

    # Style the header for top-selling categories
    for cell in ws["A5:E5"][0]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = fill_color
        cell.border = thin_border

    # Add headers for top customers
    ws.append([])  # Add an empty row for spacing
    ws.merge_cells('A7:E7')
    ws['A7'] = "Top Customers"
    ws['A7'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A7'].alignment = Alignment(horizontal="center")
    ws['A7'].fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
    ws.append(["Customer Name", "Total Orders", "Total Spend"])

    top_customers = Order.objects.values('user_id__username')\
        .annotate(total_orders=Count('serial_number'), total_spend=Sum('total_amount'))\
        .order_by('-total_spend')[:5]

    for customer in top_customers:
        ws.append([
            customer['user_id__username'],
            customer['total_orders'],
            customer['total_spend']
        ])

    # Style the header for top customers
    for cell in ws["A8:E8"][0]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = fill_color
        cell.border = thin_border

    # Add headers for orders by pincode
    ws.append([])  # Add an empty row for spacing
    ws.merge_cells('A10:E10')
    ws['A10'] = "Orders by Pincode"
    ws['A10'].font = Font(size=12, bold=True, color="FFFFFF")
    ws['A10'].alignment = Alignment(horizontal="center")
    ws['A10'].fill = PatternFill(start_color="8064A2", end_color="8064A2", fill_type="solid")
    ws.append(["Pincode", "State", "District", "Total Orders"])

    orders_by_pincode = Order.objects.values('address__pincode', 'address__state', 'address__city')\
        .annotate(total_orders=Count('serial_number'))\
        .order_by('-total_orders')[:5]

    for order in orders_by_pincode:
        ws.append([
            order['address__pincode'],
            order['address__state'],
            order['address__city'],
            order['total_orders']
        ])

    # Style the header for orders by pincode
    for cell in ws["A11:E11"][0]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = fill_color
        cell.border = thin_border

    # Prepare the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Sales_Report.xlsx'
    wb.save(response)

    return response

#----------------------------- add address in checkout ------------------------------------------------------------------------
@login_required(login_url='login')
def add_checkout_address(request):
    storage = messages.get_messages(request)
    storage.used = True

    if request.method == "POST":
        username = request.user.username
        user_data = User.objects.get(username=username)

        full_name = request.POST.get("full_name")
        phone = request.POST.get("phone")
        alt_phone = request.POST.get("alt_ph")
        pincode = request.POST.get("pincode")
        post_office = request.POST.get("post_office")
        landmark = request.POST.get("landmark")
        accessible = request.POST.get("accessible")
        address_type = request.POST.get("address_type")
        area = request.POST.get("area")
        state = request.POST.get("state")
        city = request.POST.get("city")

        check_full_name = views.name_validator(full_name)
        check_post_office = views.validate_post_office_name(post_office)
        check_pin = views.validate_pin_code(pincode)
        check_landmark = views.validate_landmark(landmark)
        check_accessible = views.validate_area(accessible)
        check_area = views.validate_area(area)
        check_city = views.validate_city(city)
        check_state = views.validate_state(state)
        check_phone = views.validate_phone(phone)
        check_alt_phone = views.validate_phone(alt_phone)

        if check_full_name[0] is True:
            messages.error(request, check_full_name[1])
            return redirect("order:checkout")
        if check_post_office[0] is True:
            messages.error(request, check_post_office[1])
            return redirect("order:checkout")
        if check_pin[0] is True:
            messages.error(request, check_pin[1])
            return redirect("order:checkout")
        if check_landmark[0] is True:
            messages.error(request, check_landmark[1])
            return redirect("order:checkout")
        if check_accessible[0] is True:
            messages.error(request, check_accessible[1])
            return redirect("order:checkout")
        if check_area[0] is True:
            messages.error(request, check_area[1])
            return redirect("order:checkout")
        if check_city[0] is True:
            messages.error(request, check_city[1])
            return redirect("order:checkout")
        if check_state[0] is True:
            messages.error(request, check_state[1])
            return redirect("order:checkout")
        if check_phone[0] is True:
            messages.error(request, check_phone[1])
            return redirect("order:checkout")
        if check_city[0] is True:
            messages.error(request, check_city[1])
            return redirect("order:checkout")
        if check_city[0] is True:
            messages.error(request, check_city[1])
            return redirect("order:checkout")
        if check_alt_phone[0] is True:
            messages.error(request, check_alt_phone[1])
            return redirect("order:checkout")

        data = UserProfile(
            user=user_data,
            pincode=pincode,
            post_office=post_office,
            landmark=landmark,
            accessible=accessible,
            address_type=address_type,
            area=area,
            state=state,
            city=city,
            alternative_phone=alt_phone,
            phone_no=phone,
            full_name=full_name,
        )
        try:
            data.full_clean()
            data.save()
            messages.success(request, "Address added")
            return redirect("order:checkout")
        except Exception as e:
            messages.error(request, f"Error! {e}")
    return render(request, "user/checkout_address_add.html")

#-------------------------------------------- edit address in checkou -------------------------------------------------------#

@login_required(login_url='login')
def edit_checkout_address(request, address_id):
    storage = messages.get_messages(request)
    storage.used = True

    addresses = UserProfile.objects.get(id=address_id)
    context = {
        "user_details": addresses,
    }
    if request.method == "POST":

        full_name = request.POST.get("full_name")
        phone = request.POST.get("phone")
        alt_phone = request.POST.get("alt_ph")
        pincode = request.POST.get("pincode")
        post_office = request.POST.get("post_office")
        landmark = request.POST.get("landmark")
        accessible = request.POST.get("accessible")
        address_type = request.POST.get("address_type")
        area = request.POST.get("area")
        state = request.POST.get("state")
        city = request.POST.get("city")

        if full_name:
            check_full_name = views.name_validator(full_name)
            if check_full_name[0] is True:
                messages.error(request, check_full_name[1])
                return redirect("order:checkout")
            addresses.full_name = full_name
        if post_office:
            check_post_office = views.validate_post_office_name(post_office)
            if check_post_office[0] is True:
                messages.error(request, check_post_office[1])
                return redirect("order:checkout")
            addresses.post_office = post_office
        if pincode:
            check_pin = views.validate_pin_code(pincode)
            if check_pin[0] is True:
                messages.error(request, check_pin[1])
                return redirect("order:checkout")
            addresses.pincode = pincode
        if landmark:
            check_landmark = views.validate_landmark(landmark)
            if check_landmark[0] is True:
                messages.error(request, check_landmark[1])
                return redirect("order:checkout")
            addresses.landmark = landmark
        if accessible:
            check_accessible = views.validate_area(accessible)
            if check_accessible[0] is True:
                messages.error(request, check_accessible[1])
                return redirect("order:checkout")
            addresses.accessible = accessible
        if area:
            check_area = views.validate_area(area)
            if check_area[0] is True:
                messages.error(request, check_area[1])
                return redirect("order:checkout")
            addresses.area = area
        if city:
            check_city = views.validate_city(city)
            if check_city[0] is True:
                messages.error(request, check_city[1])
                return redirect("order:checkout")
            addresses.city = city
        if state:
            check_state = views.validate_state(state)
            if check_state[0] is True:
                messages.error(request, check_state[1])
                return redirect("order:checkout")
            addresses.state = state
        if phone:
            check_phone = views.validate_phone(phone)
            if check_phone[0] is True:
                messages.error(request, check_phone[1])
                return redirect("order:checkout")
            addresses.phone_no = phone
        if alt_phone:
            check_alt_phone = views.validate_phone(alt_phone)
            if check_alt_phone[0] is True:
                messages.error(request, check_alt_phone[1])
                return redirect("order:checkout")
            addresses.alternative_phone = alt_phone
        if address_type:
            addresses.address_type = address_type

        try:
            addresses.save()
            messages.success(request, "Address updated")
            return redirect("order:checkout")
        except Exception as e:
            messages.error(request, f"Error ! {e}")

    return render(request, "user/checkout_address_edit.html", context)




